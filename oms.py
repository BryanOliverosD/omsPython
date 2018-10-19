import pandas as pd
from openpyxl import load_workbook
import re
from unicodedata import normalize
import math


class ShippingFalabella():
	#Atributos generales
	nombreProducto = ""
	SKU = ""
	comuna = ""
	region = ""
	nombreTienda = ""
	#Atributos necesarios para el árbol de decisión
	precio_MT = ""
	precio_BT = ""
	precio_SBT = ""
	dias_MT = ""
	dias_BT = ""
	dias_SBT = ""

	tarifaSugerida = ""

class ShippingRipley():
	#Atributos generales
	nombreProducto = ""
	SKU = ""
	comuna = ""
	region = ""
	nombreTienda = ""
	#Atributos necesarios para el árbol de decisión
	precio_MT = ""
	precio_BT = ""
	precio_SBT = ""
	dias_MT = ""
	dias_BT = ""
	dias_SBT = ""
	#Resultado árbol
	tarifaSugerida = ""

class ShippingParis():
	#Atributos generales
	nombreProducto = ""
	SKU = ""
	comuna = ""
	region = ""
	nombreTienda = ""
	#Atributos necesarios para el árbol de decisión
	precio_MT = ""
	precio_BT = ""
	precio_SBT = ""
	dias_MT = ""
	dias_BT = ""
	dias_SBT = ""
	#Resultado árbol
	tarifaSugerida = ""


##### Se copia información desde archivo shipping a propuesta #####
def CopiarHojaDetalle(name_propuesta,name_shipping):

	sheet_shipping = "Hoja 1"

	archivopropuesta = pd.ExcelFile(name_propuesta)

	excelfile_shipping = pd.read_excel(skiprows=0,io=name_shipping, sheet_name=sheet_shipping, usecols='A,B,C:I')
	excelfile_detalle = pd.read_excel(header=None,io=name_propuesta, sheet_name="Detalle", usecols='A:AA', nrows=6,index=False)
	Excelfile_comunas = pd.read_excel(skiprows=6,header=None,io=name_propuesta, sheet_name="Detalle", usecols='A',index=False)
	
	# Create a Pandas dataframe from the data.
	df = pd.DataFrame(excelfile_shipping)
	df1 = pd.DataFrame(excelfile_detalle)
	df2 = pd.DataFrame(Excelfile_comunas)

	# Se crea un nuevo archivo en el cual se pegará lo tomado desde la hoja detalles y posteriormente guardamos.
	book = load_workbook(name_propuesta)
	writer = pd.ExcelWriter(name_propuesta, engine='openpyxl')
	writer.book = book
	
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	#eliminamos la hoja detalle para que se eliminen las tablas dinámicas
	sheet_detalle = book["Detalle"]
	book.remove(sheet_detalle)
	
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	#ordenamos el excel
	df = df[['tienda','region','comuna','sku','producto','costo','dias','updated','tamaño']]
	df.to_excel(writer, sheet_name ='Shipping.csv', index=False, startcol=1)
	df1.to_excel(writer, sheet_name ='Detalle', index=False, header=False)
	df2.to_excel(writer, sheet_name ='Detalle', index=False, header=False, startcol=0, startrow=6)
	writer.save() 

## Validamos parámetros hoja base archivo propuesta actualizado ##
def ValidarParametrosBase(name_propuesta):

	sheet_base = "Base"

	excelfile = pd.read_excel(header=None,skiprows=1,io=name_propuesta, sheet_name=sheet_base, usecols='B',nrows=7)
	#posiciones variables a validar
	df = pd.DataFrame(excelfile)
	y = str(df.at[0,0])
	x = str(df.at[1,0])
	a = str(df.at[2,0])
	b = str(df.at[3,0])
	c = str(df.at[4,0])

	book = load_workbook(name_propuesta)
	#Utilizamos para poder reemplazar un valor, reescribimos todo de la misma manera salvo la variable en cuestion
	writer = pd.ExcelWriter(name_propuesta, engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

	if (y != "0.1"):
		df.at[0,0] = 0.1
	if (x != "0.0"):
		df.at[1,0] = 0
	if (a != "500.0"):
		df.at[2,0] = 500
	if (b != "750.0"):
		df.at[3,0] = 750
	if (c != "1000.0"):
		df.at[4,0] = 1000
	
	df.to_excel(writer, sheet_name='Base', index=False, startcol=1)
	writer.save()

####### Almacenar datos ##############
def AlmacenarDatos(name_propuesta):
	#creamos diccionario
	almacenador = {}
	
	excelfile = pd.read_excel(header=None,skiprows=1,io=name_propuesta, sheet_name="Shipping.csv", usecols='B,C,D,E,F,G,H,J')
	df = pd.DataFrame(excelfile)

	for fila in range(0,len(df)):
		
		if str(df.iloc[fila,2]) not in almacenador:
			#lista almacenador de objetos
			
			objetos = []
		else:

			objetos = almacenador.get(df.iloc[fila,2])

		tienda = str(df.iloc[fila,0])
		# identifica si la columna tienda es Falabella
		if tienda.lower() == "falabella":

			shipping_Falabella = ShippingFalabella()
			shipping_aux = shipping_Falabella
			shipping_Falabella.nombreTienda = df.iloc[fila,0]
			shipping_Falabella.region = df.iloc[fila,1]
			shipping_Falabella.comuna = df.iloc[fila,2]
			shipping_Falabella.SKU = df.iloc[fila,3]
			shipping_Falabella.nombreProducto = df.iloc[fila,4]
			#shippingmatrix.nombreTienda = str(df.iloc[fila,0])
			# identifica si la columna tamaño es de mt/bt/sbt
			if df.iloc[fila,7] == "MT":
				shipping_Falabella.precio_MT = df.iloc[fila,5]
				shipping_Falabella.dias_MT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "BT":
				shipping_Falabella.precio_BT = df.iloc[fila,5]
				shipping_Falabella.dias_BT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "SBT":
				shipping_Falabella.precio_SBT = df.iloc[fila,5]
				shipping_Falabella.dias_SBT = df.iloc[fila,6]

		# identifica si la columna tienda es Ripley
		elif tienda.lower() == "ripley":

			shipping_Ripley = ShippingRipley()
			shipping_aux = shipping_Ripley
			shipping_Ripley.nombreTienda = df.iloc[fila,0]
			shipping_Ripley.region = df.iloc[fila,1]
			shipping_Ripley.comuna = df.iloc[fila,2]
			shipping_Ripley.SKU = df.iloc[fila,3]
			shipping_Ripley.nombreProducto = df.iloc[fila,4]
			# identifica si la columna tamaño es de mt/bt/sbt
			if df.iloc[fila,7] == "MT":
				shipping_Ripley.precio_MT = df.iloc[fila,5]
				shipping_Ripley.dias_MT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "BT":
				shipping_Ripley.precio_BT = df.iloc[fila,5]
				shipping_Ripley.dias_BT= df.iloc[fila,6]
			elif df.iloc[fila,7] == "SBT":
				shipping_Ripley.precio_SBT = df.iloc[fila,5]
				shipping_Ripley.dias_SBT = df.iloc[fila,6]

		# identifica si la columna tienda es Paris
		elif tienda.lower() == "paris":

			shipping_Paris = ShippingParis()
			shipping_aux = shipping_Paris
			shipping_Paris.nombreTienda = df.iloc[fila,0]
			shipping_Paris.region = df.iloc[fila,1]
			shipping_Paris.comuna = df.iloc[fila,2]
			shipping_Paris.SKU = df.iloc[fila,3]
			shipping_Paris.nombreProducto = df.iloc[fila,4]
			# identifica si la columna tamaño es de mt/bt/sbt
			if df.iloc[fila,7] == "MT":
				shipping_Paris.precio_MT = df.iloc[fila,5]
				shipping_Paris.dias_MT= df.iloc[fila,6]
			elif df.iloc[fila,7] == "BT":
				shipping_Paris.precio_BT = df.iloc[fila,5]
				shipping_Paris.dias_BT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "SBT":
				shipping_Paris.precio_SBT = df.iloc[fila,5]
				shipping_Paris.dias_SBT = df.iloc[fila,6]
		# no se toma encuenta el calzado
		if df.iloc[fila,7] != "Calzado":
			objetos.append(shipping_aux)
		# Validamos si la ciudad se encuentra en el diccionario, si está se agrega a la lista que arrastra, si no se agrega al diccionario.     
		if str(df.iloc[fila,1]) not in almacenador:
			almacenador[str(df.iloc[fila,2])] = objetos
			#print(almacenador)
		else:
			almacenador[str(df.iloc[fila,2])] = objetos + almacenador[str(df.iloc[fila,2])]
	return almacenador


######## Actualizar Hoja Detalle #############
def ActualizarDetalle(name_propuesta,almacenador):

	objetos = []

	excelfile_detalle = pd.read_excel(header=None,skiprows=6,io=name_propuesta, sheet_name="Detalle", usecols='A:AA')
	df = pd.DataFrame(excelfile_detalle)
	book = load_workbook(name_propuesta)
	#Utilizamos para poder reemplazar un valor, reescribimos todo de la misma manera salvo la variable en cuestion
	writer = pd.ExcelWriter(name_propuesta, engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

	contador = 0
	# inicia recorrido de las filas, las cuales tienen el nombre de la comuna
	for fila in range(0,len(df)):

		#normalizamos las Ñ a N y transformamos a mayúsculas
		Comuna = df.at[fila,0]
		Comuna = str(Comuna).upper()
		Comuna = Comuna.replace('Ñ','N')

		#si la comuna no se encuentra en el almacenador, ya que puede que existe un problema de tildes. (normalizamos)
		if almacenador.get(Comuna) is None:
			Comuna = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", 
			normalize( "NFD", Comuna), 0, re.I)
			
			objetos = almacenador.get(Comuna)

		else:
			objetos = almacenador.get(Comuna)

		#recorremos el objeto buscando las variables correspondientes
		for i in objetos:   
			
			if i.nombreTienda.lower() == "falabella":

				if len(str(i.precio_MT)) != 0 and len(str(i.dias_MT)) != 0:
					df.at[contador,1] = i.precio_MT
					df.at[contador,2] = i.dias_MT
					
				elif len(str(i.precio_BT)) != 0 and len(str(i.dias_BT)) != 0:
					df.at[contador,7] = i.precio_BT
					df.at[contador,8] = i.dias_BT
					
				elif len(str(i.precio_SBT)) != 0 and len(str(i.dias_SBT)) != 0:
					df.at[contador,13] = i.precio_SBT
					df.at[contador,14] = i.dias_SBT
					

			elif i.nombreTienda.lower() == "ripley":

				if len(str(i.precio_MT)) != 0 and len(str(i.dias_MT)) != 0:
					df.at[contador,3] = i.precio_MT
					df.at[contador,4] = i.dias_MT
					
				elif len(str(i.precio_BT)) != 0 and len(str(i.dias_BT)) != 0:
					df.at[contador,9] = i.precio_BT
					df.at[contador,10] = i.dias_BT
					
				elif len(str(i.precio_SBT)) != 0 and len(str(i.dias_SBT)) != 0:
					df.at[contador,15] = i.precio_SBT
					df.at[contador,16] = i.dias_SBT
					

			elif i.nombreTienda.lower() == "paris":

				if len(str(i.precio_MT)) != 0 and len(str(i.dias_MT)) != 0:
					df.at[contador,5] = i.precio_MT
					df.at[contador,6] = i.dias_MT
					
				elif len(str(i.precio_BT)) != 0 and len(str(i.dias_BT)) != 0:
					df.at[contador,11] = i.precio_BT
					df.at[contador,12] = i.dias_BT
					
				elif len(str(i.precio_SBT)) != 0 and len(str(i.dias_SBT)) != 0:
					df.at[contador,17] = i.precio_SBT
					df.at[contador,18] = i.dias_SBT
					
		contador = contador + 1
	df.to_excel(writer, sheet_name='Detalle', index=False, startcol=0, startrow=6, header=None)
	writer.save()

def ArbolDecision(diasFalabella, tarifaFalabella, diasMCompetidor, tarifaMCompetidor):
	#print("DIAS F : ",diasFalabella,"Tarifa F : ",tarifaFalabella,"DIAS MC : ",diasMCompetidor, " TarifaMC : ",tarifaMCompetidor)
	valor_escenario = []
	#Se comparan los días entre Falabella y el mejor competidor
	#Más lento que el mejor competidor
	if (diasFalabella > diasMCompetidor):

		if(1 <= abs(diasMCompetidor - diasFalabella) <= 2):
			
			if(tarifaFalabella != tarifaMCompetidor):

				#print("ESCENARIO 1 - Tarifa mínima")
				tarifaMin = min(tarifaMCompetidor,tarifaFalabella)
				#print("TM : ", tarifaMin)
				valor_escenario.append("Escenario 1")
				valor_escenario.append(tarifaMin)
			else:
				#print("ESCENARIO 2 - Tarifa mínima - x%")
				tarifaMin = min(tarifaMCompetidor,tarifaFalabella)
				#print("TM : ", tarifaMin-0)
				valor_escenario.append("Escenario 2")
				valor_escenario.append(tarifaMin-0)
		
		elif (abs(diasMCompetidor - diasFalabella) > 2):
			#print("ESCENARIO 3 - Tarifa mínima - y%")
			tarifaMin = min(tarifaMCompetidor,tarifaFalabella)
			#print("TM : ", tarifaMin-10)
			valor_escenario.append("Escenario 3")
			valor_escenario.append(tarifaMin-10)

	#mismos días de entrega
	elif (diasFalabella == diasMCompetidor):

		if(tarifaFalabella != tarifaMCompetidor):
			#print("ESCENARIO 4 - Tarifa max")
			tarifaMax = max(tarifaMCompetidor,tarifaFalabella)
			#print("TM : ", tarifaMax)
			valor_escenario.append("Escenario 4")
			valor_escenario.append(tarifaMax)
		else:
			#print("ESCENARIO 5 - Tarifa max")
			tarifaMax = max(tarifaMCompetidor,tarifaFalabella)
			#print("TM : ", tarifaMax)
			valor_escenario.append("Escenario 5")
			valor_escenario.append(tarifaMax)

	# Más rápido que el mejor competidor
	elif(diasFalabella < diasMCompetidor):

		if(1 <= abs(diasMCompetidor - diasFalabella) <= 2):
			
			if(tarifaFalabella != tarifaMCompetidor):
				#print("ESCENARIO 6 - Tarifa max + a%")
				tarifaMax = max(tarifaMCompetidor,tarifaFalabella)
				#print("TM : ", tarifaMax+500)
				valor_escenario.append("Escenario 6")
				valor_escenario.append(tarifaMax+500)
			else:
				#print("ESCENARIO 7 - Tarifa max - b%")
				tarifaMax = max(tarifaMCompetidor,tarifaFalabella)
				#print("TM : ", tarifaMax+750)
				valor_escenario.append("Escenario 7")
				valor_escenario.append(tarifaMax+750)
		
		elif (abs(diasMCompetidor - diasFalabella) > 2):
			#print("ESCENARIO 8 - Tarifa max - c%")
			tarifaMax = max(tarifaMCompetidor,tarifaFalabella)
			#print("TM : ", tarifaMax + 1000)
			valor_escenario.append("Escenario 8")
			valor_escenario.append(tarifaMax+1000)
	
	return valor_escenario
############################# COMPETIDORES MT ###########################
def DefinirMejorMT(detalle):

	valor_escenario = []

	for comuna in detalle:
		#print(comuna)

		diasFalabella_MT = (detalle[comuna])[1]
		precioFalabella_MT = (detalle[comuna])[0]

		diasRipley_MT = (detalle[comuna])[3]
		precioRipley_MT = (detalle[comuna])[2]
		
		diasParis_MT = (detalle[comuna])[5]
		precioParis_MT = (detalle[comuna])[4]

		#print(diasRipley_MT,diasParis_MT)

		# Validación relacionada a la no información de Falabella
		if (diasFalabella_MT == -1 and precioFalabella_MT == -1):
			#print("SIN FALABELLA MANTENER VALOR")
			(detalle[comuna])[0] = "Sin datos"
			(detalle[comuna])[1] = "Sin datos"

			(detalle[comuna])[6] = ""
			(detalle[comuna])[7] = "Mantener valor"
			(detalle[comuna])[8] = "Mantener valor"

		# Validación relacionada a la no información de competidores	
		elif(diasRipley_MT == -1 and precioRipley_MT == -1 and precioParis_MT == -1 and diasParis_MT == -1):
			#print("Competidores Sin datos")
			#(detalle[comuna])[7] = precioFalabella_MT
			(detalle[comuna])[2] = "Sin datos"
			(detalle[comuna])[3] = "Sin datos"
			(detalle[comuna])[4] = "Sin datos"
			(detalle[comuna])[5] = "Sin datos"

			(detalle[comuna])[6] = ""
			(detalle[comuna])[7] = "Mantener valor"
			(detalle[comuna])[8] = "Mantener valor"
			
		# Validación en caso que ripley no tenga información
		elif (diasRipley_MT == -1 and precioRipley_MT == -1 and diasParis_MT != -1 and precioParis_MT != -1):
			#print("MEJOR COMPETIDOR PARIS")
			dias_mejorC = diasParis_MT
			precio_mejorC = precioParis_MT
			#árbol con Paris
			valor_escenario = ArbolDecision(diasFalabella_MT,precioFalabella_MT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[2] = "Sin datos"
			(detalle[comuna])[3] = "Sin datos"
			(detalle[comuna])[6] = valor_escenario[0]
			(detalle[comuna])[7] = valor_escenario[1]

		# Validación en caso que paris no tenga información
		elif (diasParis_MT == -1 and precioParis_MT == -1 and diasRipley_MT != -1 and precioRipley_MT != -1  ):

			dias_mejorC = diasRipley_MT
			precio_mejorC = precioRipley_MT
			#árbol con Ripley
			valor_escenario = ArbolDecision(diasFalabella_MT,precioFalabella_MT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[4] = "Sin datos"
			(detalle[comuna])[5] = "Sin datos"
			(detalle[comuna])[6] = valor_escenario[0]
			(detalle[comuna])[7] = valor_escenario[1]

		elif (diasParis_MT != -1 and diasRipley_MT != -1 and precioRipley_MT != -1 and precioParis_MT != -1):

			if(diasRipley_MT > diasParis_MT):

				dias_mejorC = diasParis_MT
				precio_mejorC = precioParis_MT
			else:
				dias_mejorC = diasRipley_MT
				precio_mejorC = precioRipley_MT

			valor_escenario = ArbolDecision(diasFalabella_MT,precioFalabella_MT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[6] = valor_escenario[0]
			(detalle[comuna])[7] = valor_escenario[1]

	return detalle

############################# COMPETIDORES BT ###########################
def DefinirMejorBT(detalle):

	valor_escenario = []

	for comuna in detalle:
		#print(comuna)

		diasFalabella_BT = (detalle[comuna])[10]
		precioFalabella_BT = (detalle[comuna])[9]

		diasRipley_BT = (detalle[comuna])[12]
		precioRipley_BT = (detalle[comuna])[11]
		
		diasParis_BT = (detalle[comuna])[14]
		precioParis_BT = (detalle[comuna])[13]

		# Validación relacionada a la no información de Falabella
		if (diasFalabella_BT == -1 and precioFalabella_BT == -1):
			#print("SIN FALABELLA MANTENER VALOR")
			(detalle[comuna])[9] = "Sin datos"
			(detalle[comuna])[10] = "Sin datos"

			(detalle[comuna])[15] = ""
			(detalle[comuna])[16] = "Mantener valor"
			(detalle[comuna])[17] = "Mantener valor"

		# Validación relacionada a la no información de competidores	
		elif(diasRipley_BT == -1 and precioRipley_BT == -1 and precioParis_BT == -1 and diasParis_BT == -1):
			#print("Competidores Sin datos")
			#(detalle[comuna])[7] = precioFalabella_BT
			(detalle[comuna])[11] = "Sin datos"
			(detalle[comuna])[12] = "Sin datos"
			(detalle[comuna])[13] = "Sin datos"
			(detalle[comuna])[14] = "Sin datos"

			(detalle[comuna])[15] = ""
			(detalle[comuna])[16] = "Mantener valor"
			(detalle[comuna])[17] = "Mantener valor"
			
		# Validación en caso que ripley no tenga información
		elif (diasRipley_BT == -1 and precioRipley_BT == -1 and diasParis_BT != -1 and precioParis_BT != -1):
			#print("MEJOR COMPETIDOR PARIS")
			dias_mejorC = diasParis_BT
			precio_mejorC = precioParis_BT
			#árbol con Paris
			valor_escenario = ArbolDecision(diasFalabella_BT,precioFalabella_BT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[12] = "Sin datos"
			(detalle[comuna])[13] = "Sin datos"
			(detalle[comuna])[15] = valor_escenario[0]
			(detalle[comuna])[16] = valor_escenario[1]

		# Validación en caso que paris no tenga información
		elif (diasParis_BT == -1 and precioParis_BT == -1 and diasRipley_BT != -1 and precioRipley_BT != -1):
			#print("MEJOR COMPETIDOR Ripley")
			dias_mejorC = diasRipley_BT
			precio_mejorC = precioRipley_BT
			#árbol con Ripley
			valor_escenario = ArbolDecision(diasFalabella_BT,precioFalabella_BT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[14] = "Sin datos"
			(detalle[comuna])[15] = "Sin datos"
			(detalle[comuna])[15] = valor_escenario[0]
			(detalle[comuna])[16] = valor_escenario[1]

		elif (diasParis_BT != -1 and diasRipley_BT != -1 and precioRipley_BT != -1 and precioParis_BT != -1):

			if(diasRipley_BT > diasParis_BT):

				dias_mejorC = diasParis_BT
				precio_mejorC = precioParis_BT
			else:
				dias_mejorC = diasRipley_BT
				precio_mejorC = precioRipley_BT

			valor_escenario = ArbolDecision(diasFalabella_BT,precioFalabella_BT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[15] = valor_escenario[0]
			(detalle[comuna])[16] = valor_escenario[1]

	return detalle

############################# COMPETIDORES SBT ###########################
def DefinirMejorSBT(detalle):

	valor_escenario = []

	for comuna in detalle:
		#print(comuna)

		diasFalabella_SBT = (detalle[comuna])[19]
		precioFalabella_SBT = (detalle[comuna])[18]

		diasRipley_SBT = (detalle[comuna])[21]
		precioRipley_SBT = (detalle[comuna])[20]
		
		diasParis_SBT = (detalle[comuna])[23]
		precioParis_SBT = (detalle[comuna])[22]

		# Validación relacionada a la no información de Falabella
		if (diasFalabella_SBT == -1 and precioFalabella_SBT == -1):
			#actualizamos valor a Sin datos
			(detalle[comuna])[18] = "Sin datos"
			(detalle[comuna])[19] = "Sin datos"

			(detalle[comuna])[24] = ""
			(detalle[comuna])[25] = "Mantener valor"
			(detalle[comuna])[26] = "Mantener valor"

		# Validación relacionada a la no información de competidores	
		elif(diasRipley_SBT == -1 and precioRipley_SBT == -1 and precioParis_SBT == -1 and diasParis_SBT == -1):
			#actualizamos valor a Sin datos
			(detalle[comuna])[20] = "Sin datos"
			(detalle[comuna])[21] = "Sin datos"
			(detalle[comuna])[22] = "Sin datos"
			(detalle[comuna])[23] = "Sin datos"

			(detalle[comuna])[24] = ""
			(detalle[comuna])[25] = "Mantener valor"
			(detalle[comuna])[26] = "Mantener valor"
			
		# Validación en caso que ripley no tenga información
		elif (diasRipley_SBT == -1 and precioRipley_SBT == -1 and diasParis_SBT != -1 and precioParis_SBT != -1):
			
			(detalle[comuna])[20] = "Sin datos"
			(detalle[comuna])[21] = "Sin datos"

			dias_mejorC = diasParis_SBT
			precio_mejorC = precioParis_SBT
			#árbol con Paris
			valor_escenario = ArbolDecision(diasFalabella_SBT,precioFalabella_SBT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[20] = "Sin datos"
			(detalle[comuna])[21] = "Sin datos"
			(detalle[comuna])[24] = valor_escenario[0]
			(detalle[comuna])[25] = valor_escenario[1]

		# Validación en caso que paris no tenga información
		elif (diasParis_SBT == -1 and precioParis_SBT == -1 and diasRipley_SBT != -1 and precioRipley_SBT != -1):
			#print("MEJOR COMPETIDOR Ripley")
			dias_mejorC = diasRipley_SBT
			precio_mejorC = precioRipley_SBT
			#árbol con Ripley
			valor_escenario = ArbolDecision(diasFalabella_SBT,precioFalabella_SBT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[22] = "Sin datos"
			(detalle[comuna])[23] = "Sin datos"
			(detalle[comuna])[24] = valor_escenario[0]
			(detalle[comuna])[25] = valor_escenario[1]

		elif (diasParis_SBT != -1 and diasRipley_SBT != -1 and precioRipley_SBT != -1 and precioParis_SBT != -1):

			if(diasRipley_SBT > diasParis_SBT):

				dias_mejorC = diasParis_SBT
				precio_mejorC = precioParis_SBT
			else:
				dias_mejorC = diasRipley_SBT
				precio_mejorC = precioRipley_SBT

			valor_escenario = ArbolDecision(diasFalabella_SBT,precioFalabella_SBT,dias_mejorC,precio_mejorC)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[24] = valor_escenario[0]
			(detalle[comuna])[25] = valor_escenario[1]

	return detalle

def ReordenarDiccionario(almacenador):

	detalle = {}

	#inicializamos matriz detalle tabla detalle
	for comuna in almacenador:

		lista_valores = [-1,-1,-1,-1,-1,-1,-1,-1,-1,
					-1,-1,-1,-1,-1,-1,-1,-1,-1,
					-1,-1,-1,-1,-1,-1,-1,-1,-1]
		detalle[comuna] = lista_valores

	#llenamos la lista de valores
	for comuna in almacenador:
		
		for objeto in almacenador[comuna]:

			#consultamos por MT en cada tienda
			if objeto.precio_MT != "":
				if (objeto.nombreTienda).lower() =="ripley":
					
					(detalle[comuna])[2] = objeto.precio_MT
					(detalle[comuna])[3] = objeto.dias_MT
					#se asignan las variables del competidor

				elif(objeto.nombreTienda).lower() == "paris":
					
					(detalle[comuna])[4] = objeto.precio_MT
					(detalle[comuna])[5] = objeto.dias_MT
					#se asignan las variables del competidor

				elif(objeto.nombreTienda).lower() == "falabella":
					
					(detalle[comuna])[0] = objeto.precio_MT
					(detalle[comuna])[1] = objeto.dias_MT

			#consultamos por BT en cada tienda
			elif objeto.precio_BT != "":
				if (objeto.nombreTienda).lower() =="ripley":
					
					(detalle[comuna])[11] = objeto.precio_BT
					(detalle[comuna])[12] = objeto.dias_BT
					#se asignan las variables del competidor

				elif(objeto.nombreTienda).lower() == "paris":
					
					(detalle[comuna])[13] = objeto.precio_BT
					(detalle[comuna])[14] = objeto.dias_BT
					#se asignan las variables del competidor

				elif(objeto.nombreTienda).lower() == "falabella":
					
					(detalle[comuna])[9] = objeto.precio_BT
					(detalle[comuna])[10] = objeto.dias_BT

			#consultamos por SBT en cada tienda
			elif objeto.precio_SBT != "":
				if (objeto.nombreTienda).lower() =="ripley":
					
					(detalle[comuna])[20] = objeto.precio_SBT
					(detalle[comuna])[21] = objeto.dias_SBT
					#se asignan las variables del competidor

				elif(objeto.nombreTienda).lower() == "paris":
					
					(detalle[comuna])[22] = objeto.precio_SBT
					(detalle[comuna])[23] = objeto.dias_SBT
					#se asignan las variables del competidor

				elif(objeto.nombreTienda).lower() == "falabella":
					
					(detalle[comuna])[18] = objeto.precio_SBT
					(detalle[comuna])[19] = objeto.dias_SBT
					
	return detalle		

def GenerarAnalisis(analisis):

	analisis = DefinirMejorMT(analisis)
	analisis = DefinirMejorBT(analisis)
	analisis = DefinirMejorSBT(analisis)

	return analisis

def GenerarResumen(analisis):

	file_name_resumen = "input/Salida_OMS.xlsx"
	contador = 0

	excelfile = pd.read_excel(header=None,skiprows=4,nrows=93,io=file_name_resumen, sheet_name="Análisis", usecols='B,E:AN',index=None)
	#posiciones variables a validar
	df = pd.DataFrame(excelfile)

	# Se crea un nuevo archivo en el cual se pegará lo tomado desde la hoja detalles y posteriormente guardamos.
	book = load_workbook(file_name_resumen)
	writer = pd.ExcelWriter(file_name_resumen, engine='openpyxl')

	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	worksheet = writer.sheets["Análisis"]

	#recorremos lo leído en el excel
	while contador < len(df):

		comuna = df.at[contador,0]
		
		datos = analisis.get(comuna)

		if analisis.get(comuna) is None:
			# normalizamos nombres comuna
			comuna = comuna.replace('Ñ','N')
			comuna = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1", 
			normalize( "NFD", comuna), 0, re.I)
			datos = analisis.get(comuna)
		
		#Escribiendo MT
		df.iloc[contador, 0] = datos[0]
		df.iloc[contador, 1] = datos[1]
		df.iloc[contador, 2] = datos[2]
		df.iloc[contador, 3] = datos[3]
		df.iloc[contador, 4] = datos[4]
		df.iloc[contador, 5] = datos[5]
		df.iloc[contador,8] = datos[6]
		df.iloc[contador,10] = datos[7]
		df.iloc[contador,11] = datos[8]
		#Escribimos BT
		df.iloc[contador,12] = datos[9]
		df.iloc[contador,13] = datos[10]
		df.iloc[contador,14] = datos[11]
		df.iloc[contador,15] = datos[12]
		df.iloc[contador,16] = datos[13]
		df.iloc[contador,17] = datos[14]
		df.iloc[contador,20] = datos[15]
		df.iloc[contador,22] = datos[16]
		df.iloc[contador,23] = datos[17]
		#Escribien SBT
		df.iloc[contador,24] = datos[18]
		df.iloc[contador,25] = datos[19]
		df.iloc[contador,26] = datos[20]
		df.iloc[contador,27] = datos[21]
		df.iloc[contador,28] = datos[22]
		df.iloc[contador,29] = datos[23]
		df.iloc[contador,32] = datos[24]
		df.iloc[contador,34] = datos[25]
		df.iloc[contador,35] = datos[26]

		contador = contador + 1
		df.to_excel(writer, sheet_name ='Análisis', index=None, header=None, startcol=4, startrow=4)
	writer.save()

def calcularNuevaTarifa(analisis):

	contador = 0

	for comuna in analisis:

		datos = analisis[comuna]

		### MT ###
		if datos[0] != datos[7]:
			datos[8] = datos[7]
		else:
			datos[8] = ""
		### BT ###
		if datos[9] != datos[16]:
			datos[17] = datos[16]
		else:
			datos[17] = ""
		### SBT ###
		if datos[18] != datos[25]:
			datos[26] = datos[25]
		else:
			datos[26] = ""

		contador = contador + 1
		
	return analisis

def RestriccionesTickets(analisis):

	for comuna in analisis:

		datos = analisis[comuna]

		print(datos[8],datos[17],datos[26])
		
		if str(datos[8]).isdigit() is True and str(datos[17]).isdigit() is True:

			if int(datos[17]) <= int(datos[8]):
				datos[17] = datos[8] + 2000

		elif str(datos[17]).isdigit() is True and str(datos[26]).isdigit() is True:
			if int(datos[17]) == int(datos[26]):
				datos[26] = datos[17] + 1000
			elif datos[26] < datos[17]:
				if datos[8] == "":
					datos[17] = datos[26]-1000
				elif datos[8] !="" and int(datos[26])-1000 > int(datos[8]):
					datos[17] = datos[26]-1000
				elif datos[8] !="" and int(datos[26])-1000 <= int(datos[8]):
					datos[26] = datos[17] + 1000
		print(datos[8],datos[17],datos[26])
	return analisis
########################################################
####################### MAIN ###########################
#########################################################

almacenador = {}
detalle = {}
analisis = {}
#parche = {}
#parche["QUILI"] = [5490, 1, 5350, 2, 3990, 1, 'Escenario 4', 5490, 'x', 9990, 4, 7850, 1, 9990, 2, 'Escenario 3', 7840, 7840, 10990, 4, 8990, 1, 9990, 2, 'Escenario 3', 8980, 5980]
file_name_shipping = "input/shipping_Falabella.xls"
file_name_propuesta = "input/propuesta2.xlsm"
#CopiarHojaDetalle(file_name_propuesta,file_name_shipping)
#ValidarParametrosBase(file_name_propuesta)
almacenador = AlmacenarDatos(file_name_propuesta)
#ActualizarDetalle(file_name_propuesta,almacenador)
detalle = ReordenarDiccionario(almacenador)
# actualizamos MT;BT;SBT
analisis = GenerarAnalisis(detalle)
analisis = calcularNuevaTarifa(analisis)
analisis = RestriccionesTickets(analisis)
"""for comuna in analisis:
	print(comuna)
	for objeto in analisis[comuna]:
		print(objeto," ",end="")
	print()
	print()"""
GenerarResumen(analisis)