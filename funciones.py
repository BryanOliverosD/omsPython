import pandas as pd
from openpyxl import load_workbook
import re
from unicodedata import normalize
import csv
import datetime
import shutil

class ShippingMatrix():
	#Atributos generales
	nombreProducto = ""
	SKU = ""
	comuna = ""
	region = ""
	nombreTienda = ""
	#Atributos necesarios para el árbol de decisión
	precioMT = ""
	precioBT = ""
	precioSBT = ""
	diasMT = ""
	diasBT = ""
	diasSBT = ""
	tarifaSugerida = ""
def leerParametros():
	global x,y,a,b,c,d
	with open('Parametros/parametros.txt') as csv_file:
		csv_reader = csv.reader(csv_file, delimiter='=')
		for row in csv_reader:
			if row[0] == "y":
				y = row[1]
			elif row[0] == "x":
				x = row[1]
			elif row[0] == "a":
				a = row[1]
			elif row[0] == "b":
				b = row[1]
			elif row[0] == "c":
				c = row[1]
			elif row[0] == "d":
				d = row[1]
				##### Se copia información desde archivo shipping a propuesta #####
def copiarHoja(nameShipping, namePropuesta):
	sheetShipping = "Hoja1"
	archivopropuesta = pd.ExcelFile(namePropuesta)
	try:
		excelfileShipping = pd.read_excel(skiprows=0,io=nameShipping, sheet_name=sheetShipping, usecols='A:I')
		#excelfileDetalle = pd.read_excel(header=None,io=namePropuesta, sheet_name="Detalle", usecols='A:AA', nrows=6,index=False)
		#Excelfile_comunas = pd.read_excel(skiprows=6,header=None,io=namePropuesta, sheet_name="Detalle", usecols='A',index=False)
		# Create a Pandas dataframe from the data.
		df = pd.DataFrame(excelfileShipping)
		#df1 = pd.DataFrame(excelfileDetalle)
		#df2 = pd.DataFrame(Excelfile_comunas)
		# Se crea un nuevo archivo en el cual se pegará lo tomado desde la hoja detalles y posteriormente guardamos.
		book = load_workbook(namePropuesta)
		writer = pd.ExcelWriter(namePropuesta, engine='openpyxl')
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		"""#eliminamos la hoja detalle para que se eliminen las tablas dinámicas
		sheet_detalle = book["Detalle"]
		book.remove(sheet_detalle)

		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)"""
		#ordenamos el excel
		df = df[['tienda','region','comuna','sku','producto','costo','dias','updated','tamano']]
		df.to_excel(writer, sheet_name = 'Hoja1', index = False, startcol = 1 )
		#df1.to_excel(writer, sheet_name ='Detalle', index=False, header=False)
		#df2.to_excel(writer, sheet_name ='Detalle', index=False, header=False, startcol=0, startrow=6)
		writer.save()
	except Exception as e:
		print(e)
## Validamos parámetros hoja base archivo propuesta actualizado ##
def validarParametros(namePropuesta):

	sheetBase = "Base"

	excelfile = pd.read_excel(header=None,skiprows=1,io=namePropuesta, sheet_name=sheetBase, usecols='B',nrows=7)
	#posiciones variables a validar
	df = pd.DataFrame(excelfile)
	y = str(df.at[0,0])
	x = str(df.at[1,0])
	a = str(df.at[2,0])
	b = str(df.at[3,0])
	c = str(df.at[4,0])

	book = load_workbook(namePropuesta)
	#Utilizamos para poder reemplazar un valor, reescribimos todo de la misma manera salvo la variable en cuestion
	writer = pd.ExcelWriter(namePropuesta, engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

	if (y != "0.1"):
		df.at[0,0] = 0.1
	if (x != "0.0"):
		df.at[1,0] = 0
	if (a != "500.0"):
		df.at[2,0] = 500
	if (b != "750.0"):
		df.at[3,0] = 750
	if (c != "1000.0"):
		df.at[4,0] = 1000

	df.to_excel(writer, sheet_name='Base', index=False, startcol=1)
	writer.save()
####### Almacenar datos ##############
def almacenarDatos(namePropuesta):
	#creamos diccionario
	almacenador = {}
	shippingAux = ShippingMatrix()
	comuna = ""

	excelfile = pd.read_excel(header=None,skiprows=1,io = namePropuesta, sheet_name="Hoja1", usecols='B,C,D,E,F,G,H,J')
	df = pd.DataFrame(excelfile)

	for fila in range(0,len(df)):
		comuna = df.iloc[fila,2]
		comuna = str(comuna).upper()
		comuna = comuna.replace('Ñ','N')
		comuna = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1",
			normalize( "NFD", comuna), 0, re.I)

		if str(comuna) not in almacenador:
			#lista almacenador de objetos
			objetos = []
		else:
			objetos = almacenador.get(comuna)
		tienda = str(df.iloc[fila,0])
		# identifica si la columna tienda es Falabella
		if tienda.lower() == "falabella":
			shippingFalabella = ShippingMatrix()
			shippingAux = shippingFalabella
			shippingFalabella.nombreTienda = df.iloc[fila,0]
			shippingFalabella.region = df.iloc[fila,1]
			shippingFalabella.comuna = comuna
			shippingFalabella.SKU = df.iloc[fila,3]
			shippingFalabella.nombreProducto = df.iloc[fila,4]
			# identifica si la columna tamaño es de mt/bt/sbt
			if df.iloc[fila,7] == "MT":
				shippingFalabella.precioMT = df.iloc[fila,5]
				shippingFalabella.diasMT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "BT":
				shippingFalabella.precioBT = df.iloc[fila,5]
				shippingFalabella.diasBT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "SBT":
				shippingFalabella.precioSBT = df.iloc[fila,5]
				shippingFalabella.diasSBT = df.iloc[fila,6]
			objetos.append(shippingFalabella)
		# identifica si la columna tienda es Ripley
		elif tienda.lower() == "ripley":
			shippingRipley = ShippingMatrix()
			shippingAux = shippingRipley
			shippingRipley.nombreTienda = df.iloc[fila,0]
			shippingRipley.region = df.iloc[fila,1]
			shippingRipley.comuna = comuna
			shippingRipley.SKU = df.iloc[fila,3]
			shippingRipley.nombreProducto = df.iloc[fila,4]
			# identifica si la columna tamaño es de mt/bt/sbt
			if df.iloc[fila,7] == "MT":
				shippingRipley.precioMT = df.iloc[fila,5]
				shippingRipley.diasMT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "BT":
				shippingRipley.precioBT = df.iloc[fila,5]
				shippingRipley.diasBT= df.iloc[fila,6]
			elif df.iloc[fila,7] == "SBT":
				shippingRipley.precioSBT = df.iloc[fila,5]
				shippingRipley.diasSBT = df.iloc[fila,6]
			objetos.append(shippingRipley)
		# identifica si la columna tienda es Paris
		elif tienda.lower() == "paris":
			shippingParis = ShippingMatrix()
			shippingAux = shippingParis
			shippingParis.nombreTienda = df.iloc[fila,0]
			shippingParis.region = df.iloc[fila,1]
			shippingParis.comuna = comuna
			shippingParis.SKU = df.iloc[fila,3]
			shippingParis.nombreProducto = df.iloc[fila,4]
			# identifica si la columna tamaño es de mt/bt/sbt
			if df.iloc[fila,7] == "MT":
				shippingParis.precioMT = df.iloc[fila,5]
				shippingParis.diasMT= df.iloc[fila,6]
			elif df.iloc[fila,7] == "BT":
				shippingParis.precioBT = df.iloc[fila,5]
				shippingParis.diasBT = df.iloc[fila,6]
			elif df.iloc[fila,7] == "SBT":
				shippingParis.precioSBT = df.iloc[fila,5]
				shippingParis.diasSBT = df.iloc[fila,6]
			objetos.append(shippingParis)
		# no se toma encuenta el calzado
		if df.iloc[fila,7] != "Calzado":
			objetos.append(shippingAux)
		# Validamos si la ciudad se encuentra en el diccionario, si está se agrega a la lista que arrastra, si no se agrega al diccionario.
		if str(df.iloc[fila,1]) not in almacenador:
			almacenador[str(comuna)] = objetos
			#print(almacenador)
		else:
			almacenador[str(comuna)] = objetos + almacenador[str(comuna)]
	return almacenador
######## Actualizar Hoja Detalle #############
def actualizarDetalle(namePropuesta,almacenador):
	objetos = []
	excelfileDetalle = pd.read_excel(header=None,skiprows=6,io=namePropuesta, sheet_name="Detalle", usecols='A:AA')
	df = pd.DataFrame(excelfileDetalle)
	book = load_workbook(namePropuesta)
	#Utilizamos para poder reemplazar un valor, reescribimos todo de la misma manera salvo la variable en cuestion
	writer = pd.ExcelWriter(namePropuesta, engine='openpyxl')
	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

	contador = 0
	# inicia recorrido de las filas, las cuales tienen el nombre de la comuna
	for fila in range(0,len(df)):

		#normalizamos las Ñ a N y transformamos a mayúsculas
		Comuna = df.at[fila,0]
		Comuna = str(Comuna).upper()
		Comuna = Comuna.replace('Ñ','N')

		#si la comuna no se encuentra en el almacenador, ya que puede que existe un problema de tildes. (normalizamos)
		if almacenador.get(Comuna) is None:
			Comuna = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1",
			normalize( "NFD", Comuna), 0, re.I)
			objetos = almacenador.get(Comuna)
		else:
			objetos = almacenador.get(Comuna)
		#recorremos el objeto buscando las variables correspondientes
		for i in objetos:
			if i.nombreTienda.lower() == "falabella":
				if len(str(i.precioMT)) != 0 and len(str(i.diasMT)) != 0:
					df.at[contador,1] = i.precioMT
					df.at[contador,2] = i.diasMT
				elif len(str(i.precioBT)) != 0 and len(str(i.diasBT)) != 0:
					df.at[contador,7] = i.precioBT
					df.at[contador,8] = i.diasBT
				elif len(str(i.precioSBT)) != 0 and len(str(i.diasSBT)) != 0:
					df.at[contador,13] = i.precioSBT
					df.at[contador,14] = i.diasSBT
			elif i.nombreTienda.lower() == "ripley":
				if len(str(i.precioMT)) != 0 and len(str(i.diasMT)) != 0:
					df.at[contador,3] = i.precioMT
					df.at[contador,4] = i.diasMT
				elif len(str(i.precioBT)) != 0 and len(str(i.diasBT)) != 0:
					df.at[contador,9] = i.precioBT
					df.at[contador,10] = i.diasBT
				elif len(str(i.precioSBT)) != 0 and len(str(i.diasSBT)) != 0:
					df.at[contador,15] = i.precioSBT
					df.at[contador,16] = i.diasSBT
			elif i.nombreTienda.lower() == "paris":
				if len(str(i.precioMT)) != 0 and len(str(i.diasMT)) != 0:
					df.at[contador,5] = i.precioMT
					df.at[contador,6] = i.diasMT
				elif len(str(i.precioBT)) != 0 and len(str(i.diasBT)) != 0:
					df.at[contador,11] = i.precioBT
					df.at[contador,12] = i.diasBT
				elif len(str(i.precioSBT)) != 0 and len(str(i.diasSBT)) != 0:
					df.at[contador,17] = i.precioSBT
					df.at[contador,18] = i.diasSBT
		contador = contador + 1
	df.to_excel(writer, sheet_name='Detalle', index=False, startcol=0, startrow=6, header=None)
	writer.save()
############################# COMPETIDORES MT ###########################
def definirMejorMT(detalle):
	valorEscenario = []
	for comuna in detalle:
		#print(comuna)
		diasFalabellaMT = (detalle[comuna])[1]
		precioFalabellaMT = (detalle[comuna])[0]
		diasRipleyMT = (detalle[comuna])[3]
		precioRipleyMT = (detalle[comuna])[2]
		diasParisMT = (detalle[comuna])[5]
		precioParisMT = (detalle[comuna])[4]
		# Validación relacionada a la no información de Falabella
		if (diasFalabellaMT == -1 and precioFalabellaMT == -1):
			#print("SIN FALABELLA MANTENER VALOR")
			(detalle[comuna])[0] = "Sin datos"
			(detalle[comuna])[1] = "Sin datos"
			(detalle[comuna])[6] = ""
			(detalle[comuna])[7] = "Mantener valor"
			(detalle[comuna])[8] = "Mantener valor"
		# Validación relacionada a la no información de competidores
		if(diasRipleyMT == -1 and precioRipleyMT == -1 and precioParisMT == -1 and diasParisMT == -1):
			#print("Competidores Sin datos")
			#(detalle[comuna])[7] = precioFalabellaMT
			(detalle[comuna])[2] = "Sin datos"
			(detalle[comuna])[3] = "Sin datos"
			(detalle[comuna])[4] = "Sin datos"
			(detalle[comuna])[5] = "Sin datos"
			(detalle[comuna])[6] = ""
			(detalle[comuna])[7] = "Mantener valor"
			(detalle[comuna])[8] = "Mantener valor"
		# Validación en caso que ripley no tenga información
		elif (diasRipleyMT == -1 and precioRipleyMT == -1 and diasParisMT != -1 and precioParisMT != -1):
			#print("MEJOR COMPETIDOR PARIS")
			diasMejorC = diasParisMT
			precioMejorC = precioParisMT
			#árbol con Paris
			valorEscenario = arbolDecision(diasFalabellaMT, 0, diasMejorC, precioMejorC, precioFalabellaMT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[2] = "Sin datos"
			(detalle[comuna])[3] = "Sin datos"
			(detalle[comuna])[6] = valorEscenario[0]
			(detalle[comuna])[7] = int(valorEscenario[1])
		# Validación en caso que paris no tenga información
		elif (diasParisMT == -1 and precioParisMT == -1 and diasRipleyMT != -1 and precioRipleyMT != -1  ):
			diasMejorC = diasRipleyMT
			precioMejorC = precioRipleyMT
			#árbol con Ripley
			valorEscenario = arbolDecision(diasFalabellaMT, 0, diasMejorC, precioMejorC, precioFalabellaMT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[4] = "Sin datos"
			(detalle[comuna])[5] = "Sin datos"
			(detalle[comuna])[6] = valorEscenario[0]
			(detalle[comuna])[7] = int(valorEscenario[1])
		elif (diasParisMT != -1 and diasRipleyMT != -1 and precioRipleyMT != -1 and precioParisMT != -1):
			if(diasRipleyMT > diasParisMT):
				diasMejorC = diasParisMT
				precioMejorC = precioParisMT
				#dias_Competidor1 =diasRipleyMT
				precioCompetidor1 = precioRipleyMT
			else:
				diasMejorC = diasRipleyMT
				precioMejorC = precioRipleyMT
				#dias_Competidor1 =diasParisMT
				precioCompetidor1 = precioParisMT
			valorEscenario = arbolDecision(diasFalabellaMT, precioCompetidor1, diasMejorC, precioMejorC, precioFalabellaMT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[6] = valorEscenario[0]
			(detalle[comuna])[7] = int(valorEscenario[1])
	return detalle
############################# COMPETIDORES BT ###########################
def definirMejorBT(detalle):
	valorEscenario = []
	for comuna in detalle:
		#print(comuna)
		diasFalabellaBT = (detalle[comuna])[10]
		precioFalabellaBT = (detalle[comuna])[9]
		diasRipleyBT = (detalle[comuna])[12]
		precioRipleyBT = (detalle[comuna])[11]
		diasParisBT = (detalle[comuna])[14]
		precioParisBT = (detalle[comuna])[13]
		# Validación relacionada a la no información de Falabella
		if (diasFalabellaBT == -1 and precioFalabellaBT == -1):
			#print("SIN FALABELLA MANTENER VALOR")
			(detalle[comuna])[9] = "Sin datos"
			(detalle[comuna])[10] = "Sin datos"
			(detalle[comuna])[15] = ""
			(detalle[comuna])[16] = "Mantener valor"
			(detalle[comuna])[17] = "Mantener valor"
		# Validación relacionada a la no información de competidores
		elif(diasRipleyBT == -1 and precioRipleyBT == -1 and precioParisBT == -1 and diasParisBT == -1):
			#print("Competidores Sin datos")
			#(detalle[comuna])[7] = precioFalabellaBT
			(detalle[comuna])[11] = "Sin datos"
			(detalle[comuna])[12] = "Sin datos"
			(detalle[comuna])[13] = "Sin datos"
			(detalle[comuna])[14] = "Sin datos"
			(detalle[comuna])[15] = ""
			(detalle[comuna])[16] = "Mantener valor"
			(detalle[comuna])[17] = "Mantener valor"
		# Validación en caso que ripley no tenga información
		elif (diasRipleyBT == -1 and precioRipleyBT == -1 and diasParisBT != -1 and precioParisBT != -1):
			#print("MEJOR COMPETIDOR PARIS")
			diasMejorC = diasParisBT
			precioMejorC = precioParisBT
			#árbol con Paris
			valorEscenario = arbolDecision(diasFalabellaBT, 0, diasMejorC, precioMejorC, precioFalabellaBT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[12] = "Sin datos"
			(detalle[comuna])[13] = "Sin datos"
			(detalle[comuna])[15] = valorEscenario[0]
			(detalle[comuna])[16] = int(valorEscenario[1])
		# Validación en caso que paris no tenga información
		elif (diasParisBT == -1 and precioParisBT == -1 and diasRipleyBT != -1 and precioRipleyBT != -1):
			#print("MEJOR COMPETIDOR Ripley")
			diasMejorC = diasRipleyBT
			precioMejorC = precioRipleyBT
			#árbol con Ripley
			valorEscenario = arbolDecision(diasFalabellaBT, 0, diasMejorC, precioMejorC, precioFalabellaBT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[14] = "Sin datos"
			(detalle[comuna])[15] = "Sin datos"
			(detalle[comuna])[15] = valorEscenario[0]
			(detalle[comuna])[16] = int(valorEscenario[1])
		elif (diasParisBT != -1 and diasRipleyBT != -1 and precioRipleyBT != -1 and precioParisBT != -1):
			if(diasRipleyBT > diasParisBT):
				diasMejorC = diasParisBT
				precioMejorC = precioParisBT
				#dias_Competidor1 =diasRipleyBT
				precioCompetidor1 = precioRipleyBT
			else:
				diasMejorC = diasRipleyBT
				precioMejorC = precioRipleyBT
				#dias_Competidor1 =diasParisBT
				precioCompetidor1 = precioParisBT
			valorEscenario = arbolDecision(diasFalabellaBT, precioCompetidor1, diasMejorC, precioMejorC, precioFalabellaBT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[15] = valorEscenario[0]
			(detalle[comuna])[16] = int(valorEscenario[1])

	return detalle
############################# COMPETIDORES SBT ###########################
def definirMejorSBT(detalle):
	valorEscenario = []
	for comuna in detalle:
		#print(comuna)
		diasFalabellaSBT = (detalle[comuna])[19]
		precioFalabellaSBT = (detalle[comuna])[18]
		diasRipleySBT = (detalle[comuna])[21]
		precioRipleySBT = (detalle[comuna])[20]
		diasParisSBT = (detalle[comuna])[23]
		precioParisSBT = (detalle[comuna])[22]
		# Validación relacionada a la no información de Falabella
		if (diasFalabellaSBT == -1 and precioFalabellaSBT == -1):
			#actualizamos valor a Sin datos
			(detalle[comuna])[18] = "Sin datos"
			(detalle[comuna])[19] = "Sin datos"
			(detalle[comuna])[24] = ""
			(detalle[comuna])[25] = "Mantener valor"
			(detalle[comuna])[26] = "Mantener valor"
		# Validación relacionada a la no información de competidores
		elif(diasRipleySBT == -1 and precioRipleySBT == -1 and precioParisSBT == -1 and diasParisSBT == -1):
			#actualizamos valor a Sin datos
			(detalle[comuna])[20] = "Sin datos"
			(detalle[comuna])[21] = "Sin datos"
			(detalle[comuna])[22] = "Sin datos"
			(detalle[comuna])[23] = "Sin datos"
			(detalle[comuna])[24] = ""
			(detalle[comuna])[25] = "Mantener valor"
			(detalle[comuna])[26] = "Mantener valor"
		# Validación en caso que ripley no tenga información
		elif (diasRipleySBT == -1 and precioRipleySBT == -1 and diasParisSBT != -1 and precioParisSBT != -1):
			(detalle[comuna])[20] = "Sin datos"
			(detalle[comuna])[21] = "Sin datos"
			diasMejorC = diasParisSBT
			precioMejorC = precioParisSBT
			#árbol con Paris
			valorEscenario = arbolDecision(diasFalabellaSBT, 0, diasMejorC, precioMejorC, precioFalabellaSBT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[20] = "Sin datos"
			(detalle[comuna])[21] = "Sin datos"
			(detalle[comuna])[24] = valorEscenario[0]
			(detalle[comuna])[25] = int(valorEscenario[1])
		# Validación en caso que paris no tenga información
		elif (diasParisSBT == -1 and precioParisSBT == -1 and diasRipleySBT != -1 and precioRipleySBT != -1):
			#print("MEJOR COMPETIDOR Ripley")
			diasMejorC = diasRipleySBT
			precioMejorC = precioRipleySBT
			#árbol con Ripley
			valorEscenario = arbolDecision(diasFalabellaSBT, 0, diasMejorC, precioMejorC, precioFalabellaSBT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[22] = "Sin datos"
			(detalle[comuna])[23] = "Sin datos"
			(detalle[comuna])[24] = valorEscenario[0]
			(detalle[comuna])[25] = int(valorEscenario[1])
		elif (diasParisSBT != -1 and diasRipleySBT != -1 and precioRipleySBT != -1 and precioParisSBT != -1):
			if(diasRipleySBT > diasParisSBT):
				diasMejorC = diasParisSBT
				precioMejorC = precioParisSBT
				#dias_Competidor1 =diasRipleySBT
				precioCompetidor1 = precioRipleySBT
			else:
				diasMejorC = diasRipleySBT
				precioMejorC = precioRipleySBT
				#dias_Competidor1 =diasParisSBT
				precioCompetidor1 = precioParisSBT
			valorEscenario = arbolDecision(diasFalabellaSBT, precioCompetidor1, diasMejorC, precioMejorC, precioFalabellaSBT)
			#insertamos el escenario y su tarifa sugerida
			(detalle[comuna])[24] = valorEscenario[0]
			(detalle[comuna])[25] = int(valorEscenario[1])
	return detalle
def arbolDecision(diasCompetidorA, tarifaCompetidorA, diasMCompetidor, tarifaMCompetidor,tarifaFalabella):
	global x,y,a,b,c
	valorEscenario = []
	#Se comparan los días entre Falabella y el mejor competidor
	#Más lento que el mejor competidor
	if (diasCompetidorA > diasMCompetidor):
		if(1 <= abs(diasMCompetidor - diasCompetidorA) <= 2):
			if(tarifaCompetidorA != tarifaMCompetidor):
				tarifaMin = min(tarifaMCompetidor,tarifaCompetidorA)
				valorEscenario.append("Escenario 1")
				if (tarifaMin == 0):
					tarifaMin = max(tarifaMCompetidor,tarifaCompetidorA)
					valorEscenario.append(tarifaMin+int((int(y)*tarifaMin/100)))
				else:
					valorEscenario.append(tarifaMin+int((int(y)*tarifaMin/100)))
			else:
				tarifaMin = min(tarifaMCompetidor,tarifaCompetidorA)
				valorEscenario.append("Escenario 2")
				if (tarifaMin == 0):
					tarifaMin = max(tarifaMCompetidor,tarifaCompetidorA)
					valorEscenario.append(tarifaMin-int(x))
				else:
					valorEscenario.append(tarifaMin-int(x))
		elif (abs(diasMCompetidor - diasCompetidorA) > 2):
			tarifaMin = min(tarifaMCompetidor,tarifaCompetidorA)
			minPorcentaje = (int(tarifaFalabella) * 85 ) / 100
			valorEscenario.append("Escenario 3")
			if (tarifaMin == 0):
					tarifaMin = max(tarifaMCompetidor,tarifaCompetidorA)
			if (tarifaMin >	minPorcentaje):
				valorEscenario.append(tarifaMin)
			else:
				valorEscenario.append(minPorcentaje)
	#mismos días de entrega
	elif (diasCompetidorA == diasMCompetidor):
		if(tarifaCompetidorA != tarifaMCompetidor):
			tarifaMax = max(tarifaMCompetidor,tarifaCompetidorA)
			valorEscenario.append("Escenario 4")
			valorEscenario.append(tarifaMax)
		else:
			tarifaMax = max(tarifaMCompetidor,tarifaCompetidorA)
			valorEscenario.append("Escenario 5")
			valorEscenario.append(tarifaMax)
	# Más rápido que el mejor competidor
	elif(diasCompetidorA < diasMCompetidor):
		if(1 <= abs(diasMCompetidor - diasCompetidorA) <= 2):
			if(tarifaCompetidorA != tarifaMCompetidor):
				tarifaMax = max(tarifaMCompetidor,tarifaCompetidorA)
				valorEscenario.append("Escenario 6")
				valorEscenario.append(tarifaMax+int(a))
			else:
				tarifaMax = max(tarifaMCompetidor,tarifaCompetidorA)
				valorEscenario.append("Escenario 7")
				valorEscenario.append(tarifaMax+int(b))
		elif (abs(diasMCompetidor - diasCompetidorA) > 2 and abs(diasMCompetidor - diasCompetidorA) < 5):
			tarifaMax = max(tarifaMCompetidor,tarifaCompetidorA)
			valorEscenario.append("Escenario 8")
			valorEscenario.append(tarifaMax+int(c))
		elif(abs(diasMCompetidor - diasCompetidorA) > 4):
			tarifaMax = max(tarifaMCompetidor,tarifaCompetidorA)
			valorEscenario.append("Escenario 9")
			valorEscenario.append(tarifaMax+int(d))
	return valorEscenario
def reordenarDiccionario(almacenador):
	detalle = {}
	#inicializamos matriz detalle tabla detalle
	for comuna in almacenador:
		lista_valores = [-1,-1,-1,-1,-1,-1,-1,-1,-1,
					-1,-1,-1,-1,-1,-1,-1,-1,-1,
					-1,-1,-1,-1,-1,-1,-1,-1,-1]
		detalle[comuna] = lista_valores
	#llenamos la lista de valores
	for comuna in almacenador:
		for objeto in almacenador[comuna]:
			#consultamos por MT en cada tienda
			if objeto.precioMT != "":
				if (objeto.nombreTienda).lower() =="ripley":
					(detalle[comuna])[2] = objeto.precioMT
					(detalle[comuna])[3] = objeto.diasMT
					#se asignan las variables del competidor
				elif(objeto.nombreTienda).lower() == "paris":
					(detalle[comuna])[4] = objeto.precioMT
					(detalle[comuna])[5] = objeto.diasMT
					#se asignan las variables del competidor
				elif(objeto.nombreTienda).lower() == "falabella":
					(detalle[comuna])[0] = objeto.precioMT
					(detalle[comuna])[1] = objeto.diasMT
			#consultamos por BT en cada tienda
			elif objeto.precioBT != "":
				if (objeto.nombreTienda).lower() =="ripley":
					(detalle[comuna])[11] = objeto.precioBT
					(detalle[comuna])[12] = objeto.diasBT
					#se asignan las variables del competidor
				elif(objeto.nombreTienda).lower() == "paris":
					(detalle[comuna])[13] = objeto.precioBT
					(detalle[comuna])[14] = objeto.diasBT
					#se asignan las variables del competidor
				elif(objeto.nombreTienda).lower() == "falabella":
					(detalle[comuna])[9] = objeto.precioBT
					(detalle[comuna])[10] = objeto.diasBT
			#consultamos por SBT en cada tienda
			elif objeto.precioSBT != "":
				if (objeto.nombreTienda).lower() =="ripley":
					(detalle[comuna])[20] = objeto.precioSBT
					(detalle[comuna])[21] = objeto.diasSBT
					#se asignan las variables del competidor
				elif(objeto.nombreTienda).lower() == "paris":
					(detalle[comuna])[22] = objeto.precioSBT
					(detalle[comuna])[23] = objeto.diasSBT
					#se asignan las variables del competidor
				elif(objeto.nombreTienda).lower() == "falabella":
					(detalle[comuna])[18] = objeto.precioSBT
					(detalle[comuna])[19] = objeto.diasSBT

	return detalle
def generarAnalisis(analisis):

	analisis = definirMejorMT(analisis)
	analisis = definirMejorBT(analisis)
	analisis = definirMejorSBT(analisis)

	return analisis
def calcularNuevaTarifa(analisis):

	contador = 0
	#lista de comunas santiago urbano
	comunas_urbanas = ["SANTIAGO","LAS CONDES","PROVIDENCIA","MAIPU","NUNOA","PUENTE ALTO","LA FLORIDA","VITACURA","PUDAHUEL","PENALOLEN",
	"SAN MIGUEL","SAN BERNARDO","QUILICURA","RECOLETA","LO BARNECHEA","LA REINA","HUECHURABA","QUINTA NORMAL","RENCA","ESTACIÓN CENTRAL","EL BOSQUE",
	"MACUL","LA CISTERNA","CONCHALI","INDEPENDENCIA","CERRO NAVIA","LA DEHESA","LA PINTANA","LO PRADO","LA GRANJA","CERRILLOS","SAN JOAQUIN","PEDRO AGUIRRE CERDA","SAN RAMON"]
	for comuna in analisis:
		datos = analisis[comuna]
		### MT ###
		# Condición para las comunas urbanas que tienen un valor constantes
		if comuna in comunas_urbanas and datos[7] != "Mantener valor" and datos[7]!="Manual":
			tarifasugerida = 3990
		else:
			tarifasugerida = datos[7]
		if datos[0] != tarifasugerida:
			datos[8] = tarifasugerida
		else:
			datos[8] = ""
		### BT ###
		if datos[9] != datos[16]:
			datos[17] = datos[16]
		else:
			datos[17] = ""
		### SBT ###
		if datos[18] != datos[25]:
			datos[26] = datos[25]
		else:
			datos[26] = ""
		contador = contador + 1
	return analisis
def aproximarValores(analisis):
	# restricción que aproxima valores a 490 o 990
	for comuna in analisis:
		datos = analisis[comuna]
		tarifaInicialMT = str(datos[7])
		tarifaInicialBT = str(datos[16])
		tarifaInicialSBT = str(datos[25])
		#### CASO MT ####
		# Preguntar por tamaño del número
		if len(tarifaInicialMT) == 4 and int(tarifaInicialMT[1:len(tarifaInicialMT)]) != 990 and int(tarifaInicialMT[1:len(tarifaInicialMT)]) != 490:
			#Se quita la primera unidad del número y se compara
			if int(tarifaInicialMT[1:len(tarifaInicialMT)]) < 501:
				datos[7] = int(str(tarifaInicialMT[0:1]) + str(490))
			else:
				datos[7] = int(str(tarifaInicialMT[0:1]) + str(990))
		#Se quitan las primeras 2 unidades del número y se compara
		elif len(tarifaInicialMT) == 5 and int(tarifaInicialMT[2:len(tarifaInicialMT)]) != 990 and int(tarifaInicialMT[2:len(tarifaInicialMT)]) != 490:
			if int(tarifaInicialMT[2:len(tarifaInicialMT)]) < 501:
				datos[7] = int(str(tarifaInicialMT[0:2]) + str(490))
			else:
				datos[7] = int(str(tarifaInicialMT[0:2]) + str(990))
		#Se quitan las primeras 3 unidades del número y se compara
		elif len(tarifaInicialMT) == 6 and tarifaInicialMT != "Manual" and int(tarifaInicialMT[3:len(tarifaInicialMT)]) != 990 and int(tarifaInicialMT[3:len(tarifaInicialMT)]) != 490:
			if int(tarifaInicialMT[3:len(tarifaInicialMT)]) < 501:
				datos[7] = int(str(tarifaInicialMT[0:3]) + str(490))
			else:
				datos[7] = int(str(tarifaInicialMT[0:3]) + str(990))
		### CASO BT ###
		if len(tarifaInicialBT) == 4 and int(tarifaInicialBT[1:len(tarifaInicialBT)]) != 990 and int(tarifaInicialBT[1:len(tarifaInicialBT)]) != 490:
			#Se quita la primera unidad del número y se compara
			if int(tarifaInicialBT[1:len(tarifaInicialBT)]) < 501:
				datos[16] = int(str(tarifaInicialBT[0:1]) + str(490))
			else:
				datos[16] = int(str(tarifaInicialBT[0:1]) + str(990))
		#Se quitan las primeras 2 unidades del número y se compara
		elif len(tarifaInicialBT) == 5 and int(tarifaInicialBT[2:len(tarifaInicialBT)]) != 990 and int(tarifaInicialBT[2:len(tarifaInicialBT)]) != 490:
			if int(tarifaInicialBT[2:len(tarifaInicialBT)]) < 501:
				datos[16] = int(str(tarifaInicialBT[0:2]) + str(490))
			else:
				datos[16] = int(str(tarifaInicialBT[0:2]) + str(990))
		#Se quitan las primeras 3 unidades del número y se compara
		elif tarifaInicialBT != "Manual" and len(tarifaInicialBT) == 6 and int(tarifaInicialBT[3:len(tarifaInicialBT)]) != 990 and int(tarifaInicialBT[3:len(tarifaInicialBT)]) != 490:
			if int(tarifaInicialBT[3:len(tarifaInicialBT)]) < 501:
				datos[16] = int(str(tarifaInicialBT[0:3]) + str(490))
			else:
				datos[16] = int(str(tarifaInicialBT[0:3]) + str(990))
		### CASO SBT ###
		if len(tarifaInicialSBT) == 4 and int(tarifaInicialSBT[1:len(tarifaInicialSBT)]) != 990 and int(tarifaInicialSBT[1:len(tarifaInicialSBT)]) != 490:
			#Se quita la primera unidad del número y se compara
			if int(tarifaInicialSBT[1:len(tarifaInicialSBT)]) < 501:
				datos[25] = int(str(tarifaInicialSBT[0:1]) + str(490))
			else:
				datos[25] = int(str(tarifaInicialSBT[0:1]) + str(990))
		#Se quitan las primeras 2 unidades del número y se compara
		elif len(tarifaInicialSBT) == 5 and int(tarifaInicialSBT[2:len(tarifaInicialSBT)]) != 990 and int(tarifaInicialSBT[2:len(tarifaInicialSBT)]) != 490:
			if int(tarifaInicialSBT[2:len(tarifaInicialSBT)]) < 501:
				datos[25] = int(str(tarifaInicialSBT[0:2]) + str(490))
			else:
				datos[25] = int(str(tarifaInicialSBT[0:2]) + str(990))
		#Se quitan las primeras 3 unidades del número y se compara
		elif len(tarifaInicialSBT) == 6 and tarifaInicialSBT != "Manual" and int(tarifaInicialSBT[3:len(tarifaInicialSBT)]) != 990 and int(tarifaInicialSBT[3:len(tarifaInicialSBT)]) != 490:
			if int(tarifaInicialSBT[3:len(tarifaInicialSBT)]) < 501:
				datos[25] = int(str(tarifaInicialSBT[0:3]) + str(490))
			else:
				datos[25] = int(str(tarifaInicialSBT[0:3]) + str(990))

		"""if len(tarifaInicialMT) == 4 and tarifaInicialMT[1:len(tarifaInicialMT)] != 990 and tarifaInicialMT[1:len(tarifaInicial)] != 490:
			#quitamos primera cifra
			tarifa = tarifaInicial[1:len(tarifaInicial)]
			#se calcula la diferencia positiva
			tarifa1 = abs(990 - int(tarifa))
			tarifa2 = abs(490 - int(tarifa))
			#se calcula el valor minimo
			tarifaMin = min(tarifa1,tarifa2)
			if tarifaMin == tarifa1:
				datos[7] = int(str(tarifaInicial[0:1]) + str(990))
			else:
				datos[7] = int(str(tarifaInicial[0:1]) + str(490))

		elif len(tarifaInicial) == 5 and tarifaInicial[2:len(tarifaInicial)] != 990 and tarifaInicial[2:len(tarifaInicial)] != 490:
			#quitamos primeras 2 cifras
			tarifa = tarifaInicial[2:len(tarifaInicial)]
			#se calcula la diferencia positiva
			tarifa1 = abs(990 - int(tarifa))
			tarifa2 = abs(490 - int(tarifa))
			#se calcula el valor minimo
			tarifaMin = min(tarifa1,tarifa2)
			if tarifaMin == tarifa1:
				datos[7] = int(str(tarifaInicial[0:2]) + str(990))
			else:
				datos[7] = int(str(tarifaInicial[0:2]) + str(490))
			#datos[7] = tarifaMin + int(tarifaInicial)

		elif len(tarifaInicial) == 6 and tarifaInicial[3:len(tarifaInicial)] != 990 and tarifaInicial[3:len(tarifaInicial)] != 490:
			#quitamos primeras 3 cifras
			tarifa = tarifaInicial[3:len(tarifaInicial)]
			t_mayor = abs(990 - int(tarifa))
			tMenor = abs(490 - int(tarifa))
			#se calcula el valor minimo
			tarifaMin = min(tarifa1,tarifa2)
			if tarifaMin == tarifa1:
				datos[7] = int(str(tarifaInicial[0:3]) + str(990))
			else:
				datos[7] = int(str(tarifaInicial[0:3]) + str(490))"""
		analisis[comuna] = datos
	return analisis
def restriccionesTickets(analisis):
	for comuna in analisis:
		datos = analisis[comuna]
		if str(datos[8]).isdigit() is True and str(datos[17]).isdigit() is True:
			if int(datos[17]) <= int(datos[8]):
				datos[17] = datos[8] + 2000
		if str(datos[17]).isdigit() is True and str(datos[26]).isdigit() is True:
			if int(datos[17]) == int(datos[26]):
				datos[26] = datos[17] + 1000
			elif datos[26] < datos[17]:
				if datos[8] == "":
					datos[17] = datos[26]-500
				elif datos[8] !="" and int(datos[26])-500 > int(datos[8]):
					datos[17] = datos[26]-500
				elif datos[8] !="" and int(datos[26])-500 <= int(datos[8]):
					datos[26] = datos[17] + 500
	return analisis
def generarResumen(analisis):
	now=datetime.datetime.now().strftime("%d%m%y_%H%M%S")
	fileNameResumen = "Plantilla/Salida_OMS.xlsx"
	fileNameSalida = "Output/Salida_OMS_"+now+".xlsx"
	shutil.copy(fileNameResumen,fileNameSalida)

	contador = 0

	excelfile = pd.read_excel(header=None,skiprows=4,nrows=93,io=fileNameSalida, sheet_name="Análisis", usecols='B,E:AN',index=None)
	#posiciones variables a validar
	df = pd.DataFrame(excelfile)

	# Se crea un nuevo archivo en el cual se pegará lo tomado desde la hoja detalles y posteriormente guardamos.
	book = load_workbook(fileNameSalida)
	writer = pd.ExcelWriter(fileNameSalida, engine='openpyxl')

	writer.book = book
	writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
	worksheet = writer.sheets["Análisis"]
	#recorremos lo leído en el excel
	while contador < len(df):
		comuna = df.at[contador,0]
		datos = analisis.get(comuna)
		if analisis.get(comuna) is None:
			# normalizamos nombres comuna
			comuna = comuna.replace('Ñ','N')
			comuna = re.sub(r"([^n\u0300-\u036f]|n(?!\u0303(?![\u0300-\u036f])))[\u0300-\u036f]+", r"\1",
			normalize( "NFD", comuna), 0, re.I)
			datos = analisis.get(comuna)





		#Escribiendo MT
		df.iloc[contador, 0] = datos[0]
		df.iloc[contador, 1] = datos[1]
		df.iloc[contador, 2] = datos[2]
		df.iloc[contador, 3] = datos[3]
		df.iloc[contador, 4] = datos[4]
		df.iloc[contador, 5] = datos[5]
		df.iloc[contador,8] = datos[6]
		df.iloc[contador,10] = datos[7]
		df.iloc[contador,11] = datos[8]

		#Escribimos BT
		df.iloc[contador,12] = datos[9]
		df.iloc[contador,13] = datos[10]
		df.iloc[contador,14] = datos[11]
		df.iloc[contador,15] = datos[12]
		df.iloc[contador,16] = datos[13]
		df.iloc[contador,17] = datos[14]
		df.iloc[contador,20] = datos[15]
		if comuna.lower() == "iquique" or comuna.lower() == "antofagasta":
			df.iloc[contador,22] = 21990
			df.iloc[contador,23] = ""
			#df.iloc[contador,23] = 21990
		elif comuna.lower() == "calama":
			df.iloc[contador,22] = 31990
			df.iloc[contador,23] = ""
			#df.iloc[contador,23] = 31990
		elif comuna.lower() == "punta arenas":
			df.iloc[contador,22] = 39990
			df.iloc[contador,23] = ""
			#df.iloc[contador,23] = 39990
		else:
			df.iloc[contador,22] = datos[16]
			df.iloc[contador,23] = datos[17]
		#Escribien SBT
		df.iloc[contador,24] = datos[18]
		df.iloc[contador,25] = datos[19]
		df.iloc[contador,26] = datos[20]
		df.iloc[contador,27] = datos[21]
		df.iloc[contador,28] = datos[22]
		df.iloc[contador,29] = datos[23]
		df.iloc[contador,32] = datos[24]
		if comuna.lower() == "iquique" or comuna.lower() == "antofagasta":
			df.iloc[contador,34] = 29990
			df.iloc[contador,35] = ""
			#df.iloc[contador,35] = 29990
		elif comuna.lower() == "calama":
			df.iloc[contador,34] = 41990
			df.iloc[contador,35] = ""
			#df.iloc[contador,35] = 41990
		elif comuna.lower() == "punta arenas":
			df.iloc[contador,34] = 79990
			df.iloc[contador,35] = ""
			#df.iloc[contador,35] = 79990
		else:
			df.iloc[contador,34] = datos[25]
			df.iloc[contador,35] = datos[26]

		contador = contador + 1
		df.to_excel(writer, sheet_name ='Análisis', index=None, header=None, startcol=4, startrow=4)
	writer.save()
#def porcentaje4Bandas(nameBandas):

